from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date
from typing import List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.repositories.reports_repo import ReportRow


def _fmt_dmy(d: date) -> str:
    return f"{d.day:02d}/{d.month:02d}/{d.year}"


def _auto_fit(ws) -> None:
    for col in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col)
        for row in range(1, ws.max_row + 1):
            v = ws.cell(row=row, column=col).value
            if v is None:
                continue
            max_len = max(max_len, len(str(v)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


_sheet_bad = re.compile(r"[\[\]\:\*\?\/\\]+")


def _safe_sheet_name(name: str) -> str:
    s = name.strip()
    s = _sheet_bad.sub("-", s)
    return s[:31] if len(s) > 31 else s


@dataclass(frozen=True)
class ReportsMetadata:
    date_from: date
    date_to: date
    client_name: str


class ReportsExcelExporter:
    HEADERS = [
        "CONSECUTIVO",
        "FECHA CONSECUTIVO",
        "PATRONO",
        "CORREO",
        "COLABORADOR",
        "FECHA INCIDENTE",
        "FALTA",
        "ESTADO",
        "OBSERVACIONES",
        "REVISIÓN",
    ]

    @staticmethod
    def build_workbook(*, rows: List[ReportRow], meta: ReportsMetadata) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = _safe_sheet_name("Reporte")

        # Header block
        ws.append(["Reporte de incidencias disciplinarias (UD)"])
        ws.append(["Rango (received_day):", _fmt_dmy(meta.date_from), "a", _fmt_dmy(meta.date_to)])
        ws.append(["Patrono:", meta.client_name])
        ws.append([])

        ws.append(ReportsExcelExporter.HEADERS)

        for r in rows:
            ws.append(
                [
                    r.consecutivo,
                    _fmt_dmy(r.fecha_consecutivo),
                    r.patrono,
                    r.correo,
                    r.colaborador,
                    _fmt_dmy(r.fecha_incidente),
                    r.falta,
                    r.estado,
                    r.observaciones,
                    r.revision,
                ]
            )

        _auto_fit(ws)
        return wb